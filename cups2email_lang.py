#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys

import configdb
from ooop import OOOP
from openpyxl import load_workbook


def cups2email(template=None, filename=None):
    if not filename:
        sys.exit("No filename defined")
    if not template:
        sys.exit("No template defined")
    wb = load_workbook(filename)
    ws = wb.active
    O = OOOP(**configdb.ooop)
    pol_obj = O.GiscedataPolissa
    cups_obj = O.GiscedataCupsPs
    templ_obj = O.PoweremailTemplates
    templ_id = templ_obj.search([('name', '=', template)])
    templ_info = templ_obj.read(templ_id)

    # Skip header row
    row = 2
    while ws.cell(row=row, column=1).value:
        cups = ws.cell(row=row, column=1).value
        print cups
        try:
            cups_id = cups_obj.search([('name', '=', cups)])[0]
        except:
            print "El CUPS no existe"
            row += 1
            continue
        try:
            pol_id = pol_obj.search([('cups', '=', cups_id)])[0]
            print pol_id
        except:
            print "No hay pólizas activas asociadas al CUPS"
            row += 1
            continue
        ctx = {
            'active_ids': [pol_id],
            'active_id': pol_id,
            'template_id': templ_id,
            'src_model': 'giscedata.polissa',
            'src_rec_ids': [pol_id],
            'from': templ_info[0]['enforce_from_account'][0]
        }
        params = {
            'state': 'single',
            'priority': 0,
            'from': ctx['from']
        }
        try:
            wz_id = O.PoweremailSendWizard.create(params, ctx)
            O.PoweremailSendWizard.send_mail([wz_id], ctx)
        except Exception:
            raise
        row += 1
